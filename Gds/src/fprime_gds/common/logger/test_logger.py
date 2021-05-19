"""
test_logger.py:

A wrapper on the openpyxl library that provides the ability to log messages in an excel sheet. Used
by the test API to log events, asserts, test cases and user messages. The documentation for openpyxl
can be found here:
https://openpyxl.readthedocs.io/en/stable/index.html

If the openpyxl library isn't installed, this class does nothing.

This class uses a write-only optimization that should allow for creating large log files without
hogging too much memory. Write-only optimization can be found here:
https://openpyxl.readthedocs.io/en/stable/optimized.html#write-only-mode

:author: koran
"""
import datetime
import os
import threading
import time

# If openpyxl isn't installed, ignore all functionality in this module
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.utils.exceptions import WorkbookAlreadySaved

    MODULE_INSTALLED = True
except ImportError:
    MODULE_INSTALLED = False


class TestLogger:
    """
    User-accessible colors. Can be used for the color arguments
    """

    __test__ = False
    BROWN = "E6CCB3"
    RED = "FF9999"
    ORANGE = "FFCC99"
    YELLOW = "FFFF99"
    GREEN = "ADEBAD"
    BLUE = "99CCFF"
    PURPLE = "CC99FF"
    GRAY = "D9D9D9"
    WHITE = "FFFFFF"

    """
    User-accessible styles. Must be used for the style arguments
    """
    BOLD = "BOLD"
    ITALICS = "ITALICS"
    UNDERLINED = "UNDERLINED"

    if MODULE_INSTALLED:
        __align = Alignment(vertical="top", wrap_text=True)
        __font_name = "calibri"
        __time_fmt = "%H:%M:%S.%f"

    def __init__(self, output_path, time_format=None, font_name=None):
        """
        Constructs a TestLogger

        Args:
            output_path: a path where log files will
            time_format: an optional string to specify the timestamp format. See datetime.strftime
            font_name: an optional string to specify the font
        """
        if not MODULE_INSTALLED:
            return

        if not isinstance(output_path, str):
            raise TypeError(
                "Test Logger requires a filename where the output can be saved."
            )
        self.start_time = time.time()
        date_string = datetime.datetime.fromtimestamp(self.start_time).strftime(
            "%Y-%m-%dT%H:%M:%S"
        )
        self.filename = os.path.join(output_path, "TestLog_{}.xlsx".format(date_string))
        self.workbook = Workbook(write_only=True)
        self.worksheet = self.workbook.create_sheet()
        self.ws_saved = False

        if time_format is None:
            self.time_format = self.__time_fmt
        else:
            self.time_format = time_format

        if font_name is None:
            self.font_name = self.__font_name
        else:
            self.font_name = font_name

        timestring = datetime.datetime.fromtimestamp(self.start_time).strftime(
            self.time_format
        )
        self.worksheet.column_dimensions["A"].width = len(timestring) + 1
        self.worksheet.column_dimensions["D"].width = 120

        top = []
        date_string = datetime.datetime.fromtimestamp(self.start_time).strftime(
            "%H:%M:%S.%f on %m/%d/%Y"
        )
        top.append(self.__get_cell("Test began at " + date_string))
        self.worksheet.append(top)

        labels = ["Log Time", "Case ID", "Sender", "Message"]
        header = self.__get_ws_row(labels, style=self.BOLD)
        self.worksheet.append(header)

        self.case_id = "NA"

        self.lock = threading.Lock()

    def log_message(self, message, sender="NA", color=None, style=None, case_id=None):
        """
        Logs a message to the TestLog. Each message will include a timestamp beforehand.
        Note: Once specified, the test case's case_id will persist in the logs until it is
        specified again.

        Args:
            message: a message to log (str).
            sender: a short string describing who created the log message
            color: a string object containing a color hex code "######"
            style: a string choosing 1 of 3 formatting options (ITALICS, BOLD, UNDERLINED)
            case_id: a short identifier to denote which test case the log message belongs to
        """
        if not MODULE_INSTALLED:
            return

        ts = time.time()
        timestring = datetime.datetime.fromtimestamp(ts).strftime(self.time_format)

        if case_id is not None:
            if not isinstance(case_id, str):
                case_id = str(case_id)
            self.case_id = case_id

        strings = [timestring, self.case_id, sender, message]
        self.lock.acquire()
        try:
            print("{} [{}] {}".format(timestring, sender, message))
            if not self.ws_saved:
                row = self.__get_ws_row(strings, color, style)
                self.worksheet.append(row)
        except WorkbookAlreadySaved:
            self.ws_saved = True
            print(
                "{} [{}] {}".format(
                    timestring, "TestLogger", "Workbook has already been saved."
                )
            )
        finally:
            self.lock.release()

    def close_log(self):
        """
        Saves the write-only workbook. Should be called only once when the log is completed.
        """
        if not MODULE_INSTALLED:
            return

        self.workbook.save(filename=self.filename)
        self.ws_saved = True

    def __get_cell(self, string, color=None, style=None):
        """
        Helper method for log message. This method takes a string as well as color and style
        arguments to create a write-only cell.

        Args:
            string: a string object to be written to the cell's value field.
            color: a string object containing a color hex code "######"
            style: a string choosing 1 of 3 formatting options (ITALICS, BOLD, UNDERLINED)
        """
        if not MODULE_INSTALLED:
            return

        cell = WriteOnlyCell(self.worksheet, value=string)
        if color is not None:
            # pylint: disable=E0237
            cell.fill = PatternFill("solid", fgColor=color)
        # pylint: disable=E0237
        cell.font = Font(
            name=self.font_name,
            bold=(style == self.BOLD),
            italic=(style == self.ITALICS),
            underline=("single" if style == self.UNDERLINED else "none"),
        )
        # pylint: disable=E0237
        cell.alignment = self.__align
        return cell

    def __get_ws_row(self, strings, color=None, style=None):
        row = []
        for string in strings:
            row.append(self.__get_cell(string, color, style))
        return row
