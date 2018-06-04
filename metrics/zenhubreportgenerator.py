import sys
import requests
import datetime
import time
import openpyxl
import openpyxl.cell
import openpyxl.chart
import openpyxl.chart.axis
import openpyxl.styles
import openpyxl.styles.colors
import openpyxl.styles.fills
import openpyxl.utils
import openpyxl.worksheet.copier
import shutil
import reportconfig

BOARD_SUFFIX = "/board"
EVENTS_SUFFIX = "/events"
ISSUES_SUFFIX = "/issues"
MILESTONES_SUFFIX = "/milestones"

HEADER = ["Repo", "Number", "Name", "Estimate", "Percent Complete", "Milestone", "Actual Completion Date",
          "Original Plan Completion Date", "Planned Completion Date", "Proposed Completion Date", "", "Table Dates",
          "Planned Progress", "Proposed Progress", "Actual Progress"]


RED_FILL = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
GREEN_FILL = openpyxl.styles.PatternFill(start_color='FF00BB00', end_color='FF00BB00', fill_type='solid')
BLUE_FILL = openpyxl.styles.PatternFill(start_color='FF8888FF', end_color='FF8888FF', fill_type='solid')
CLEAR_FILL = openpyxl.styles.PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FF000000')


def get_column_number(column):
    number = 0
    for letter in column:
        number = number * 26 + ord(letter) - ord('A') + 1
    return number


def open_workbook(filename):
    workbook = openpyxl.load_workbook(filename=filename)
    return workbook


def copy_worksheet(workbook, source, title=str(datetime.date.today()), index=0):
    # find most recent worksheet, copy it to a new one, and clear all cell coloring
    target = workbook.create_sheet(title, index)
    openpyxl.worksheet.copier.WorksheetCopy(source, target).copy_worksheet()
    return target


def update_line_item(worksheet, line, row):
    column = 1
    for entry in line:
        cell = worksheet.cell(column=column, row=row)
        if (not isinstance(cell.value, datetime.datetime) and cell.value != entry) \
                or (isinstance(cell.value, datetime.datetime) and cell.value.date() != entry):
            if (column != HEADER.index("Percent Complete") + 1) or cell.value < entry:
                cell.value = entry
                cell.fill = GREEN_FILL
        else:
            cell.fill = CLEAR_FILL
        column = column + 1


def add_line_item(worksheet, line, row, fill=CLEAR_FILL):
    column = 1
    for value in line:
        cell = worksheet.cell(column=column, row=row)
        cell.value = value
        cell.fill = fill
        column = column + 1


def generate_issue_line(pipeline_weights, repo, issue_json):
    estimate = issue_json.get("estimate")
    if estimate is not None:
        estimate = int(estimate.get("value"))
    weight = 0
    pipeline = issue_json.get("pipeline")
    if pipeline is not None:
        if pipeline_weights.get(issue_json.get("pipeline").get("name")) is not None:
            weight = int(pipeline_weights.get(issue_json.get("pipeline").get("name")))
    milestone = issue_json.get("milestone")
    if milestone is not None:
        milestone = milestone.get("title")
    completion_date = issue_json.get("closed_at")
    if completion_date is not None:
        completion_date = datetime.datetime.strptime(completion_date, "%Y-%m-%dT%H:%M:%SZ").date()
    line = [repo, issue_json.get("number"), issue_json.get("title"), estimate, weight, milestone, completion_date]
    return line


def add_data_point(worksheet, estimate_col='D', weight_col='E', report_date_col='L', actual_val_col='O'):
    # add data point to graph, maybe update referenced named range
    current_value = 0.0
    for row in range(2, len(worksheet['A']) + 1):
        estimate = worksheet[estimate_col + str(row)]
        if estimate.value is None or estimate.value == "":
            continue
        weight = worksheet[weight_col + str(row)]
        if weight.value is None or weight.value == "":
            continue
        current_value = current_value + float(estimate.value * weight.value / 100.0)
    row = 2
    new_cell = worksheet[report_date_col + str(row)]
    while new_cell.value is not None and new_cell.value != "":
        if new_cell.value == datetime.date.today():
            break
        row = row + 1
        new_cell = worksheet[report_date_col + str(row)]
    new_cell.value = datetime.date.today()
    worksheet[actual_val_col + str(row)].value = current_value


def add_sum_datapoint(worksheet, estimate_col='D', weight_col='E', report_date_col='L', actual_val_col='O'):
    size = str(max([len(worksheet[estimate_col]), len(worksheet[weight_col])]))
    row = 2
    new_cell = worksheet[report_date_col + str(row)]
    while new_cell.value is not None and new_cell.value != "":
        if new_cell.value == datetime.date.today():
            break
        row = row + 1
        new_cell = worksheet[report_date_col + str(row)]
    new_cell.value = datetime.date.today()
    worksheet[actual_val_col + str(row)].value = ("=SUMPRODUCT(" + estimate_col + "$2:" + estimate_col + "$" + size + ", "
                                                  + weight_col + "$2:" + weight_col + "$" + size + ") / 100")


def store_sum_datapoint(worksheet, estimate_col='D', weight_col='E', report_date_col='L', actual_val_col='O'):
    # make calculated sum data static
    current_value = 0.0
    for row in range(2, len(worksheet['A']) + 1):
        estimate = worksheet[estimate_col + str(row)]
        if estimate.value is None or estimate.value == "":
            continue
        weight = worksheet[weight_col + str(row)]
        if weight.value is None or weight.value == "":
            continue
        current_value = current_value + float(estimate.value * weight.value / 100.0)
    row = 2
    date_cell = worksheet[report_date_col + str(row)]
    while date_cell.value is not None and date_cell.value != "":
        store_cell = worksheet[actual_val_col + str(row)]
        if str(store_cell.value).startswith("=SUMPRODUCT("):
            store_cell.value = current_value
        row = row + 1
        date_cell = worksheet[report_date_col + str(row)]


def calculate_plan(worksheet, estimate_col='D', plan_date_col='I', report_date_col='L', report_plan_val_col='M', offset_plan_date_col = None):
    plan = {}
    for row in range(2, len(worksheet[report_plan_val_col]) + 1):
        worksheet[report_plan_val_col + str(row)].value = None
    for issue_row in range(2, len(worksheet['A']) + 1):
        issue_plan = worksheet[plan_date_col + str(issue_row)]
        if issue_plan is None or issue_plan.value is None or plan.get(issue_plan.value.date()) is not None:
            continue
        plan[issue_plan.value.date()] = None
    for plan_date in sorted(plan.keys()):
        plan_row = 2
        table_plan_date = worksheet[report_date_col + str(plan_row)]
        while table_plan_date.value is not None and table_plan_date.value != "":
            if table_plan_date.value.date() == plan_date:
                break
            plan_row = plan_row + 1
            table_plan_date = worksheet[report_date_col + str(plan_row)]
        table_plan_date.value = plan_date
        plan_date_array = plan_date_col + "$2:" + plan_date_col + "$" + str(len(worksheet[plan_date_col]))
        estimate_array = estimate_col + "$2:" + estimate_col + "$" + str(len(worksheet[estimate_col]))
        sum_function = "=SUMIF(" + plan_date_array + ",\"<=\"&" + report_date_col + str(plan_row) + ", " + estimate_array + ")"
        if offset_plan_date_col is not None:
            offset_date_array = offset_plan_date_col + "$2:" + offset_plan_date_col + "$" + str(len(worksheet[offset_plan_date_col]))
            sum_function = sum_function + " + SUMIF(" + offset_date_array + ",\"<=\"&MIN(" + report_date_col + str(plan_row) + "," + plan_date_array + ")," + estimate_array + ")"
        worksheet[report_plan_val_col + str(plan_row)].value = sum_function


def calculate_replan(worksheet):
    calculate_plan(worksheet, plan_date_col='J', report_plan_val_col='N', offset_plan_date_col='I')


def create_chart(workbook, worksheet, date_col='L', actual_val_col='O', plan_val_col='M', replan_val_col='N', sheet=False):
    # create a chart in a new worksheet to display the new data
    chart = openpyxl.chart.LineChart()
    chart.title = "Progress Report"
    # chart.style = 12
    chart.y_axis.title = "Earned Progress"

    chart.y_axis.crossAx = 500
    chart.x_axis = openpyxl.chart.axis.DateAxis(crossAx=100)
    chart.x_axis.number_format = 'd-mmm'
    chart.x_axis.baseTimeUnit = "days"
    chart.x_axis.majorUnit = 7
    chart.x_axis.majorTimeUnit = "days"
    chart.x_axis.minorUnit = 1
    chart.x_axis.minorTimeUnit = "days"
    chart.x_axis.minorTickMark = "in"
    chart.x_axis.title = "Timeline"
    chart.legend.position = 'b'
    chart.width = chart.width * 2
    chart.height = chart.height * 2

    max_row = len(worksheet[date_col])
    date_labels = openpyxl.chart.Reference(worksheet, min_row=2, max_row=max_row, min_col=get_column_number(date_col))
    plan_data = openpyxl.chart.Reference(worksheet, min_row=1, max_row=max_row, min_col=get_column_number(plan_val_col))
    replan_data = openpyxl.chart.Reference(worksheet, min_row=1, max_row=max_row, min_col=get_column_number(replan_val_col))
    actual_data = openpyxl.chart.Reference(worksheet, min_row=1, max_row=max_row, min_col=get_column_number(actual_val_col))

    chart.add_data(plan_data, titles_from_data=True)
    chart.add_data(replan_data, titles_from_data=True)
    chart.add_data(actual_data, titles_from_data=True)

    color_list = ["BB0000", "00BB00", "0000BB"]
    for index in range(3):
        chart.series[index].graphicalProperties.line.solidFill = color_list[index]
        chart.series[index].graphicalProperties.line.width = 28600
        chart.series[index].marker.symbol = "circle"
        chart.series[index].marker.size = 5
        chart.series[index].marker.graphicalProperties.solidFill = color_list[index]
        chart.series[index].marker.graphicalProperties.line.solidFill = color_list[index]
    chart.set_categories(date_labels)

    if sheet:
        chartsheet = workbook['Chart']
        if chartsheet is None:
            chartsheet = workbook.create_chartsheet()
        chartsheet.add_chart(chart)
    else:
        row = 2
        cell = worksheet['A' + str(row)]
        cell2 = worksheet['A' + str(row + 1)]
        cell3 = worksheet['A' + str(row + 2)]
        while cell3.value is not None or cell2.value is not None or cell.value is not None:
            row = row + 1
            cell = worksheet['A' + str(row)]
            cell2 = worksheet['A' + str(row + 1)]
            cell3 = worksheet['A' + str(row + 2)]
        worksheet.add_chart(chart, "A" + str(row + 3))


def update_worksheet(options, worksheet, append=False):
    # call update line item a lot, then add data point, and finally create chartsheet
    for row in range(2, len(worksheet['A']) + 1):
        starting_cell = worksheet.cell(column=1, row=row)
        if starting_cell.value is not None and starting_cell.value in options.git_repo_list:
            issue = get_issue_for_row(options, worksheet, row)
            if issue is None:
                starting_cell.fill = RED_FILL
                worksheet.cell(column=2, row=row).fill = RED_FILL
                worksheet.cell(column=3, row=row).fill = RED_FILL
                continue
            line = generate_issue_line(options, starting_cell.value, issue)
            update_line_item(worksheet, line, row)
        else:
            for col in range(1, 7):
                worksheet.cell(column=col, row=row).fill = CLEAR_FILL
    if append:
        append_issues(options, worksheet)


def append_issues(options, worksheet):
    max_seen = {}
    row = 2
    current_repo_cell = worksheet['A' + str(row)]
    current_issue_num_cell = worksheet['B' + str(row)]
    while current_repo_cell.value is not None and current_repo_cell.value is not "":
        if max_seen.get(current_repo_cell.value) is None or max_seen.get(current_repo_cell.value) < current_issue_num_cell.value:
            max_seen[current_repo_cell.value] = current_issue_num_cell.value
        row = row + 1
        current_repo_cell = worksheet['A' + str(row)]
        current_issue_num_cell = worksheet['B' + str(row)]
    options.issues = get_all_issues(options, limits=max_seen)
    for repo in options.git_repo_list:
        repo_max = max_seen.get(repo)
        if repo_max is None:
            repo_max = 0
        issue_numbers = sorted(options.issues[repo].keys())
        for issue_number in issue_numbers:
            if issue_number <= repo_max:
                continue
            issue = options["issues"][repo][issue_number]
            line = generate_issue_line(options.pipeline_weights, repo, issue)
            add_line_item(worksheet, line, row, fill=GREEN_FILL)
            row = row + 1


def get_all_issues(options, issues={}, limits={}):
    for git_repo in options["git_repo_list"]:
        git_repo_url = options["git_base_url"] + '/' + git_repo
        git_repo_id = options.get("git_repo_ids").get(git_repo)
        if git_repo_id is None:
            git_repo_data = requests.get(git_repo_url, headers=options["git_auth_header"])
            git_repo_id = str(git_repo_data.json()["id"])
            options["git_repo_ids"][git_repo] = git_repo_id
        git_issues_url = git_repo_url + ISSUES_SUFFIX
        get_next = True
        while get_next:
            get_next = False
            git_issues_data = requests.get(git_issues_url, headers=options["git_auth_header"], params=options["git_issue_params"])
            link_header = git_issues_data.headers.get("Link")
            if link_header is not None:
                links = link_header.split(" ")
                if links[1] == "rel=\"next\",":
                    next_link = links[0]
                    git_issues_url = next_link[1:len(next_link) - 2]
                    get_next = True
            git_issues_json = git_issues_data.json()
            # iterate through all of the issues retrieved
            if issues.get(git_repo) is None:
                issues[git_repo] = {}
            for git_issue_json in git_issues_json:
                git_number = git_issue_json.get("number")
                if limits.get(git_repo) is not None and limits.get(git_repo) >= git_number:
                    get_next = False
                    break
                label_dicts = git_issue_json.get("labels")
                labels = [label_dict.get("name") for label_dict in label_dicts]
                if "*" not in options.get("included_labels") and not any(label in options.get("included_labels") for label in labels):
                    continue
                if any(label in options.get("excluded_labels") for label in labels):
                    continue
                zen_issue_api_url = options["zen_base_url"] + '/' + git_repo_id + ISSUES_SUFFIX + "/" + str(git_number)
                zen_issue_data = requests.get(zen_issue_api_url, headers=options["zen_auth_header"])
                while zen_issue_data.status_code == 403 and zen_issue_data.text == u'{"message":"API Rate limit reached."}':
                    print("waiting for zenhub api reset")
                    time.sleep(60.5)
                    print("resuming")
                    zen_issue_data = requests.get(zen_issue_api_url, headers=options["zen_auth_header"])
                zen_issue_json = zen_issue_data.json()
                git_issue_json.update(zen_issue_json)
                issues[git_repo][git_number] = git_issue_json
    return issues


def get_issue(options, git_repo, number, name):
    git_repo_url = options["git_base_url"] + '/' + git_repo
    git_repo_id = options.get("git_repo_ids").get(git_repo)
    if git_repo_id is None:
        git_repo_data = requests.get(git_repo_url, headers=options["git_auth_header"])
        git_repo_id = str(git_repo_data.json()["id"])
        options["git_repo_ids"][git_repo] = git_repo_id
    if number is None:
        if name is None:
            return None
        if options.get("issues") is None:
            options["issues"] = get_all_issues(options)
        for repo in options.get("issues").keys():
            for number in options.get("issues").get(repo).keys():
                if options.get("issues").get(repo).get(number).get("title") == name:
                    return options.get("issues").get(repo).get(number)
        return None
    git_issue_url = git_repo_url + ISSUES_SUFFIX + "/" + str(number)
    git_issue_data = requests.get(git_issue_url, headers=options["git_auth_header"], params=options["git_issue_params"])
    git_issue_json = git_issue_data.json()
    zen_issue_api_url = options["zen_base_url"] + '/' + git_repo_id + ISSUES_SUFFIX + "/" + str(number)
    zen_issue_data = requests.get(zen_issue_api_url, headers=options["zen_auth_header"])
    while zen_issue_data.status_code == 403 and zen_issue_data.text == u'{"message":"API Rate limit reached."}':
        print("waiting for zenhub api reset")
        time.sleep(60.5)
        print("resuming")
        zen_issue_data = requests.get(zen_issue_api_url, headers=options["zen_auth_header"])
    zen_issue_json = zen_issue_data.json()
    git_issue_json.update(zen_issue_json)
    return git_issue_json


def get_issue_for_row(options, worksheet, row):
    repo = worksheet.cell(column=1, row=row).value
    number = worksheet.cell(column=2, row=row).value
    name = worksheet.cell(column=3, row=row).value
    return get_issue(options, repo, number, name)


def create_workbook(options):
    # get all of the issues, dump them into a workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = str(datetime.date.today())
    for index in range(len(HEADER)):
        cell = worksheet.cell(column=index + 1, row=1, value=HEADER[index])
        cell.fill = BLUE_FILL
        cell.font = openpyxl.styles.Font(bold=True)
    options["issues"] = get_all_issues(options)
    row = 2
    for repo in options["git_repo_list"]:
        issue_numbers = sorted(options["issues"][repo].keys())
        for issue_number in issue_numbers:
            issue = options["issues"][repo][issue_number]
            line = generate_issue_line(options, repo, issue)
            add_line_item(worksheet, line, row)
            row = row + 1
    workbook.save(options["target_workbook_filename"])
    return workbook


def update_workbook(options, workbook=None, append=False):
    if workbook is None:
        shutil.copy(options["source_workbook_filename"], options["target_workbook_filename"])
        workbook = open_workbook(options["target_workbook_filename"])
    new_worksheet = workbook.active
    new_worksheet.title = str(datetime.date.today())
    update_worksheet(options, new_worksheet, append=append)
    calculate_plan(new_worksheet)
    calculate_replan(new_worksheet)
    store_sum_datapoint(new_worksheet)
    add_sum_datapoint(new_worksheet)
    create_chart(workbook, new_worksheet)
    workbook.save(options["target_workbook_filename"])
    return workbook


def main(args):
    if len(args) < 4:
        print("Usage: python zenhubreportgenerator.py <create|update|append> <existing_file> <config_file>")
        return
    config = reportconfig.ReportConfiguration(args[3])
    config.source_workbook_filename = args[2]
    config.target_workbook_filename = config.planning_report_filename + "-" + str(datetime.date.today()) + ".xlsx"
    if args[1] == "create":
        workbook = create_workbook(config)
    if args[1] == "append":
        workbook = update_workbook(config, append=True)
    elif args[1] == "update":
        workbook = update_workbook(config)


if __name__ == '__main__':
    main(sys.argv)
